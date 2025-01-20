import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.implicitly_wait(5)

# Download the file
driver.find_element(By.ID, "downloadButton").click()
time.sleep(3)
file_Path = "C:/Users/shubh/Downloads/download.xlsx"


# Open, Update and Save the file
def update_excel_data(filePath, fruitName, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for j in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=j).value == colName:
            Dict['col'] = j

    for i in range(2, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == fruitName:
                Dict['row'] = i

    sheet.cell(row=Dict['row'], column=Dict['col']).value = new_value
    book.save(file_Path)


update_excel_data(file_Path, "Apple", "price", "10000")

# Upload the file
inputFile = driver.find_element(By.XPATH, "//input[@type='file']")
inputFile.send_keys("C:/Users/shubh/Downloads/download.xlsx")

successLocator = (By.XPATH, "//div[@class='Toastify']/div/div/div/div[2]")
wait = WebDriverWait(driver, 5)
wait.until(expected_conditions.visibility_of_element_located(successLocator))
print(driver.find_element(*successLocator).text)

